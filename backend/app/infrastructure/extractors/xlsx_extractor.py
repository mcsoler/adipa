import io

from openpyxl import load_workbook

from app.interfaces.extractors.text_extractor import ITextExtractor


class XlsxExtractor(ITextExtractor):
    """Extrae texto de archivos XLSX usando openpyxl (SRP: solo extrae XLSX)."""

    def extract(self, file_bytes: bytes) -> str:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        rows: list[str] = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                cells = [str(cell) for cell in row if cell is not None and str(cell).strip()]
                if cells:
                    rows.append(" | ".join(cells))
        return "\n".join(rows).strip()
